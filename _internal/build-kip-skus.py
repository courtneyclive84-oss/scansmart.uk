#!/usr/bin/env python3
"""
SCANSMART PWA SQLite BUILD — kip-skus.sqlite (~5,076 rows, offline barcode lookup)

WHY THIS EXISTS:
The PWA at app.scansmart.uk needs a client-side SQLite cache for offline scans.
This script combines the canonical I500 corpus (129 rows marked 'Verified' in
the workbook) with a freshness-ranked top-up from the OFF dump pull (up to
4,947 rows filtered by data completeness) and writes a single ~1MB SQLite file
that ships in the PWA bundle.

LOCKED SPEC (28 May 2026; I500 count corrected 29 May 2026 — see I500 NOTE):
  Source        = 129 I500 Verified + up to 4,947 from OFF dump (nutrient-rich)
  Ranking       = OFF rows filtered by data completeness, then by freshness
  Library       = sql.js (loads the .sqlite file in the PWA, queries in memory)
  Distribution  = ship as static asset in PWA bundle, precached by service worker
  Schema        = 18 columns minimal render-shape (see CREATE TABLE below)
  Lookup integ. = SQLite-first, OFF API as fallback (the PWA's lookupProduct())
  Build         = this script, runs on demand on Ras's Mac
  Refresh       = manual, anchored to PWA deploys

I500 NOTE (29 May 2026, founder decision): build from the 129 rows whose
Barcode Status == 'Verified'. The 148 figure from §103 is the strict-methodology
AUDIT denominator (138/148 = 93.2% absence), NOT the count of Verified rows —
the two are different concepts. The 19 remaining 'TBC – needs scanning' rows
that carry an OFF assessment are blocked on real data fixes (5 have placeholder
barcodes 'Unknown - rescan'; 2 duplicate-barcode pairs) and stay TBC until
rescanned. Revisit 148 after the rescans; do not flip status to force the count.

WHAT YOU DO:
    cd ~/Documents/ScanSmart/scansmart-site/_internal
    python3 build-kip-skus.py

DEPENDENCIES:
    Python 3.x + openpyxl (pip install openpyxl)
    sqlite3 is in Python stdlib — no separate install
    No network access required (uses local files only)

INPUTS:
    ~/Documents/ScanSmart/i500/KiP_i500_Consolidated_v1_2026-04-22.xlsx
        Product Catalogue sheet, headers at Excel row 4, data rows 5-onwards.
        Filter to Barcode Status == 'Verified' → ~148 rows.

    ~/Documents/ScanSmart/scansmart-site/_internal/universe-pwa-sqlite-LATEST.json
        Output of pull-from-dump-for-pwa-sqlite.py — nutrient-rich UNIVERSE.
        If this file doesn't exist yet, falls back to universe-LATEST.json
        (no nutrients — dry-run only; produces an I500-only SQLite with no
        OFF top-up because the completeness filter rejects all nutrient-less
        rows). Run pull-from-dump-for-pwa-sqlite.py on your Mac to enable the
        full ~5,076-row build.

OUTPUT:
    ~/Documents/ScanSmart/KiP/Web/DEPLOY_v4.10+game+website_2026-05-12/kip-skus.sqlite

VERIFICATION:
    After running, the script prints a self-audit: total rows, sources split,
    completeness stats, sample query against the resulting DB.
"""
import json
import os
import sys
import sqlite3
from datetime import date, datetime

# ============================================================================
# CONFIG
# ============================================================================

I500_WORKBOOK = os.path.expanduser(
    "~/Documents/ScanSmart/i500/KiP_i500_Consolidated_v1_2026-04-22.xlsx"
)
UNIVERSE_NUTRIENT_RICH = os.path.expanduser(
    "~/Documents/ScanSmart/scansmart-site/_internal/universe-pwa-sqlite-LATEST.json"
)
UNIVERSE_FALLBACK = os.path.expanduser(
    "~/Documents/ScanSmart/scansmart-site/_internal/universe-LATEST.json"
)
OUTPUT_SQLITE = os.path.expanduser(
    "~/Documents/ScanSmart/KiP/Web/DEPLOY_v4.10+game+website_2026-05-12/kip-skus.sqlite"
)

OFF_TARGET_COUNT = 4947   # top-up size from OFF UNIVERSE (per locked spec)

# The 18-column SQLite schema (locked spec)
CREATE_SCHEMA = """
CREATE TABLE IF NOT EXISTS skus (
    barcode             TEXT PRIMARY KEY,
    product_name        TEXT NOT NULL,
    brand               TEXT,
    category            TEXT,
    nutrition_grade     TEXT,
    nova_group          INTEGER,
    salt_100g           REAL,
    sugars_100g         REAL,
    fat_100g            REAL,
    saturated_fat_100g  REAL,
    fiber_100g          REAL,
    protein_100g        REAL,
    carbohydrates_100g  REAL,
    energy_kcal_100g    REAL,
    quantity            TEXT,
    source              TEXT NOT NULL,
    i500_pid            INTEGER,
    last_modified       TEXT
);
-- barcode is PRIMARY KEY → automatically indexed
CREATE INDEX IF NOT EXISTS idx_skus_source ON skus(source);
"""

# Fields required for an OFF row to pass the completeness filter.
#
# RELAXED 29 May 2026 to match OFF data reality. Original spec required all
# 8 per-100g nutrients + grade + NOVA + category (11 fields). Empirical probe
# of 200 dump records showed only ~70% have full nutriments; the rest have
# grade computed but the canonical nutrients dict is empty (private-label
# foreign imports). The dump-pull's extract_nutrients fallback recovers
# nutrients from nutriscore_data.components for those records, but only
# the 5 Nutri-Score-relevant nutrients (salt, sugars, saturated_fat, energy,
# fiber). Fat-total, protein, carbohydrates aren't in nutriscore_data — they
# stay None for the 30% slice.
#
# So the build filter requires only the verdict-critical fields the PWA
# renderer absolutely needs to draw the headline card:
#   - nutrition_grade (the A-E letter)
#   - nova_group (the 1-4 number)
#   - category (so the card knows what it's looking at)
#   - salt_100g + sugars_100g (the two FSA traffic-light pillars + the
#                              two the I500 workbook also carries)
#   - saturated_fat_100g + energy_kcal_100g (the other two I500 fields)
#
# The other 4 nutrients (fat total, protein, carbs, fiber) are nice-to-have
# but not blocking. The PWA renderer shows "data unavailable" on missing
# fields per the §10 clinical-safety rule.
COMPLETENESS_REQUIRED = [
    "ns",                      # nutrition_grade — empty string fails
    "nv",                      # nova_group — None fails
    "cat",                     # category — empty string fails
    "salt_100g",
    "sugars_100g",
    "saturated_fat_100g",
    "energy_kcal_100g",
]


# ============================================================================
# HELPERS
# ============================================================================

def _str_or_none(v, max_len=None):
    """Coerce to str, strip, return None if empty. Optional max_len truncate."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def _float_or_none(v):
    """Coerce to float, return None if not numeric or NaN."""
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:    # NaN check
            return None
        return f
    except (TypeError, ValueError):
        return None


def _int_or_none(v):
    """Coerce to int, return None if not numeric."""
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


# ============================================================================
# I500 WORKBOOK READER
# ============================================================================

def read_i500_workbook(path):
    """Read the I500 workbook's Product Catalogue sheet.
    Returns list of dicts in the SQLite schema shape.
    Filters to rows where Barcode Status == 'Verified'."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    if not os.path.exists(path):
        sys.exit(f"ERROR: I500 workbook not found at {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Product Catalogue" not in wb.sheetnames:
        sys.exit("ERROR: 'Product Catalogue' sheet not found in workbook")
    ws = wb["Product Catalogue"]

    # Header row is Excel row 4 (index 3 in 0-indexed list)
    # Data starts at Excel row 5
    rows = list(ws.iter_rows(min_row=5, values_only=True))

    out = []
    skipped_unverified = 0
    skipped_no_barcode = 0
    skipped_no_name = 0
    skipped_duplicate = 0
    seen_barcodes = set()

    for row in rows:
        if not row or all(c is None for c in row):
            continue
        # Column layout (verified by probing 29 May 2026):
        #   0: ID, 1: Barcode (EAN), 2: Barcode Status, 3: Product Name, 4: Brand,
        #   5: Category, 6: Sugar (g/100g), 7: Salt (g/100g), 8: Sat Fat (g/100g),
        #   9: Calories (kcal/100g), 10: Sugar RAG, 11: Salt RAG, 12: Serving (g),
        #   13: Sugar (tsp), 14: Salt (sachets), 15: Variants, 16: OFF, 17: Source Scope
        ident          = _str_or_none(row[0])
        barcode        = _str_or_none(row[1])
        barcode_status = _str_or_none(row[2])
        product_name   = _str_or_none(row[3], 120)
        brand          = _str_or_none(row[4], 60)
        category       = _str_or_none(row[5], 60)
        sugars         = _float_or_none(row[6])
        salt           = _float_or_none(row[7])
        sat_fat        = _float_or_none(row[8])
        calories       = _float_or_none(row[9])

        # Only include rows explicitly marked Verified
        if barcode_status != "Verified":
            skipped_unverified += 1
            continue
        if not barcode:
            skipped_no_barcode += 1
            continue
        if not product_name:
            skipped_no_name += 1
            continue
        if barcode in seen_barcodes:
            skipped_duplicate += 1
            continue
        seen_barcodes.add(barcode)

        # Extract i500_pid from the "I500-001" ident format
        i500_pid = None
        if ident and ident.startswith("I500-"):
            i500_pid = _int_or_none(ident[5:])

        out.append({
            "barcode": barcode,
            "product_name": product_name,
            "brand": brand,
            "category": category,
            "nutrition_grade": None,            # workbook doesn't have a letter grade
            "nova_group": None,                  # workbook doesn't have NOVA group
            "salt_100g": salt,
            "sugars_100g": sugars,
            "fat_100g": None,                    # workbook only has Sat Fat, not total Fat
            "saturated_fat_100g": sat_fat,
            "fiber_100g": None,                  # workbook doesn't have fibre
            "protein_100g": None,                # workbook doesn't have protein
            "carbohydrates_100g": None,          # workbook doesn't have carbs
            "energy_kcal_100g": calories,
            "quantity": None,                    # workbook doesn't carry product quantity
            "source": "i500",
            "i500_pid": i500_pid,
            "last_modified": None,               # workbook generation date is per-workbook, not per-row
        })

    print(f"  I500 workbook: {len(out)} verified rows loaded "
          f"(skipped: {skipped_unverified} unverified, {skipped_no_barcode} no-barcode, "
          f"{skipped_no_name} no-name, {skipped_duplicate} duplicate)")
    return out


# ============================================================================
# UNIVERSE JSON READER (nutrient-rich variant) + FALLBACK
# ============================================================================

def read_universe_json():
    """Read the nutrient-rich UNIVERSE if it exists; fall back to the
    metadata-only one with a clear warning if not."""
    if os.path.exists(UNIVERSE_NUTRIENT_RICH):
        path = UNIVERSE_NUTRIENT_RICH
        is_rich = True
    elif os.path.exists(UNIVERSE_FALLBACK):
        path = UNIVERSE_FALLBACK
        is_rich = False
    else:
        sys.exit(f"ERROR: Neither {UNIVERSE_NUTRIENT_RICH} nor {UNIVERSE_FALLBACK} exists. "
                 f"Run pull-from-dump-for-pwa-sqlite.py on your Mac first.")

    with open(path) as f:
        data = json.load(f)
    print(f"  UNIVERSE source: {path}")
    print(f"  UNIVERSE records: {len(data):,}")
    if not is_rich:
        print(f"  ⚠  Fallback in use — this UNIVERSE has no nutrient values.")
        print(f"  ⚠  Completeness filter will reject all OFF rows; output will be I500-only.")
        print(f"  ⚠  Run pull-from-dump-for-pwa-sqlite.py for a real ~5,076-row build.")
    return data, is_rich


# ============================================================================
# COMPLETENESS FILTER + FRESHNESS RANK
# ============================================================================

def is_complete(rec):
    """A UNIVERSE record passes completeness if every COMPLETENESS_REQUIRED
    field is populated (non-None, non-empty)."""
    for field in COMPLETENESS_REQUIRED:
        v = rec.get(field)
        if v is None:
            return False
        if isinstance(v, str) and not v.strip():
            return False
    return True


def freshness_key(rec):
    """Sort key for freshness ranking. Newer last_modified first; missing dates last."""
    lm = rec.get("last_modified")
    if not lm:
        return ""
    return lm   # ISO date strings sort lexicographically the same as chronologically


def select_off_top_up(universe, exclude_barcodes, target_count):
    """Filter UNIVERSE for completeness, exclude any barcodes already in the I500
    set, sort by freshness (descending), return up to target_count rows."""
    # Stage 1: filter for completeness
    complete = [r for r in universe if is_complete(r)]
    pre_dedup = len(complete)

    # Stage 2: exclude barcodes already in I500 (so the source column stays honest)
    complete = [r for r in complete if r.get("c") not in exclude_barcodes]
    post_dedup = len(complete)
    dedup_removed = pre_dedup - post_dedup

    # Stage 3: sort by freshness, take top N
    complete.sort(key=freshness_key, reverse=True)
    top = complete[:target_count]

    print(f"  UNIVERSE complete-filtered: {pre_dedup:,}")
    print(f"  UNIVERSE after I500-barcode dedup: {post_dedup:,} (removed {dedup_removed})")
    print(f"  UNIVERSE top-{target_count} by freshness: {len(top):,}")
    return top


def universe_to_schema(rec):
    """Convert a UNIVERSE record (compact field names) into a SQLite schema row dict."""
    return {
        "barcode": rec.get("c"),
        "product_name": rec.get("n"),
        "brand": rec.get("b"),
        "category": rec.get("cat"),
        "nutrition_grade": (rec.get("ns") or "").strip() or None,
        "nova_group": rec.get("nv"),
        "salt_100g": rec.get("salt_100g"),
        "sugars_100g": rec.get("sugars_100g"),
        "fat_100g": rec.get("fat_100g"),
        "saturated_fat_100g": rec.get("saturated_fat_100g"),
        "fiber_100g": rec.get("fiber_100g"),
        "protein_100g": rec.get("protein_100g"),
        "carbohydrates_100g": rec.get("carbohydrates_100g"),
        "energy_kcal_100g": rec.get("energy_kcal_100g"),
        "quantity": rec.get("q"),
        "source": "off",
        "i500_pid": None,
        "last_modified": rec.get("last_modified"),
    }


# ============================================================================
# SQLITE WRITER
# ============================================================================

INSERT_SQL = """
INSERT INTO skus (
    barcode, product_name, brand, category, nutrition_grade, nova_group,
    salt_100g, sugars_100g, fat_100g, saturated_fat_100g,
    fiber_100g, protein_100g, carbohydrates_100g, energy_kcal_100g,
    quantity, source, i500_pid, last_modified
) VALUES (?, ?, ?, ?, ?, ?,  ?, ?, ?, ?,  ?, ?, ?, ?,  ?, ?, ?, ?)
"""

SCHEMA_KEYS = [
    "barcode", "product_name", "brand", "category", "nutrition_grade", "nova_group",
    "salt_100g", "sugars_100g", "fat_100g", "saturated_fat_100g",
    "fiber_100g", "protein_100g", "carbohydrates_100g", "energy_kcal_100g",
    "quantity", "source", "i500_pid", "last_modified"
]


def write_sqlite(rows, out_path):
    """Create a fresh kip-skus.sqlite at out_path, insert every row in `rows`."""
    # Ensure the output directory exists
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    # Remove any previous version — full rebuild every time
    if os.path.exists(out_path):
        os.remove(out_path)
    conn = sqlite3.connect(out_path)
    cur = conn.cursor()
    cur.executescript(CREATE_SCHEMA)
    # Bulk insert via executemany
    values = [tuple(r.get(k) for k in SCHEMA_KEYS) for r in rows]
    cur.executemany(INSERT_SQL, values)
    conn.commit()
    cur.execute("VACUUM")
    cur.execute("ANALYZE")
    conn.commit()
    conn.close()


# ============================================================================
# SELF-AUDIT
# ============================================================================

def audit_sqlite(out_path):
    """Open the resulting database and print stats so the build is verifiable
    without leaving the script."""
    size_bytes = os.path.getsize(out_path)
    conn = sqlite3.connect(out_path)
    cur = conn.cursor()
    total = cur.execute("SELECT COUNT(*) FROM skus").fetchone()[0]
    by_source = cur.execute(
        "SELECT source, COUNT(*) FROM skus GROUP BY source"
    ).fetchall()
    fully_complete = cur.execute("""
        SELECT COUNT(*) FROM skus
        WHERE nutrition_grade IS NOT NULL AND nova_group IS NOT NULL
        AND salt_100g IS NOT NULL AND sugars_100g IS NOT NULL
        AND fat_100g IS NOT NULL AND saturated_fat_100g IS NOT NULL
        AND fiber_100g IS NOT NULL AND protein_100g IS NOT NULL
        AND carbohydrates_100g IS NOT NULL AND energy_kcal_100g IS NOT NULL
    """).fetchone()[0]
    sample_lookups = []
    for code in ["5000157024671", "5029788174890", "5449000000996"]:
        r = cur.execute(
            "SELECT barcode, product_name, brand, source FROM skus WHERE barcode = ?",
            (code,)
        ).fetchone()
        sample_lookups.append((code, r))
    conn.close()
    print()
    print("─" * 70)
    print(f"BUILD VERIFICATION ({out_path}, {size_bytes:,} bytes)")
    print("─" * 70)
    print(f"  Total rows: {total:,}")
    for src, n in by_source:
        print(f"    source='{src}': {n:,} rows")
    print(f"  Fully nutrient-complete rows: {fully_complete:,} ({fully_complete/total*100:.1f}%)")
    print()
    print("  Sample barcode lookups:")
    for code, r in sample_lookups:
        if r:
            print(f"    {code} → {r[1]!r} [{r[2]}] (source: {r[3]})")
        else:
            print(f"    {code} → NOT FOUND in kip-skus.sqlite")
    print()


# ============================================================================
# MAIN
# ============================================================================

def main():
    print(f"PWA SQLite BUILD — kip-skus.sqlite")
    print(f"Output: {OUTPUT_SQLITE}")
    print()

    print("─" * 70)
    print("Stage 1: load I500 verified rows from workbook")
    print("─" * 70)
    i500_rows = read_i500_workbook(I500_WORKBOOK)
    i500_barcodes = {r["barcode"] for r in i500_rows}
    print()

    print("─" * 70)
    print("Stage 2: load UNIVERSE JSON + apply completeness filter + freshness rank")
    print("─" * 70)
    universe, is_rich = read_universe_json()
    off_top = select_off_top_up(universe, i500_barcodes, OFF_TARGET_COUNT)
    off_rows = [universe_to_schema(r) for r in off_top]
    print()

    print("─" * 70)
    print("Stage 3: write SQLite database")
    print("─" * 70)
    all_rows = i500_rows + off_rows
    print(f"  Total to insert: {len(all_rows):,} "
          f"({len(i500_rows):,} I500 + {len(off_rows):,} OFF top-up)")
    write_sqlite(all_rows, OUTPUT_SQLITE)
    print(f"  ✓ Wrote {OUTPUT_SQLITE}")

    audit_sqlite(OUTPUT_SQLITE)

    if not is_rich:
        print("⚠  DRY-RUN ONLY: UNIVERSE had no nutrients.")
        print("⚠  Run pull-from-dump-for-pwa-sqlite.py on your Mac for the full ~5,076-row build.")


if __name__ == "__main__":
    main()
